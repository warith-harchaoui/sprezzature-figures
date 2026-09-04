"""
Tests for sprezzature_figures.studio.ingest: the CSV, XLSX, and clipboard
readers, semantic type detection (guessing whether a column holds a date, a
category, or a number, from its values rather than its header name), and
the profiler, which summarizes a loaded table before the user picks a
chart.

Author
------
Warith Harchaoui <warith.harchaoui@gmail.com>
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
import pytest

from sprezzature_figures.studio.ingest import (
    clipboard_fingerprint,
    clipboard_warnings,
    csv_fingerprint,
    csv_warnings,
    excel_fingerprint,
    excel_warnings,
    list_sheets,
    parse_clipboard_text,
    preview_csv,
    profile_dataframe,
    read_csv,
    read_excel,
    sniff_csv,
    validate_upload_size,
)
from sprezzature_figures.studio.ingest.profiler import MAX_UPLOAD_BYTES
from sprezzature_figures.studio.ingest.semantic_types import detect_semantic_type

# --------------------------------------------------------------------------
# fixtures
# --------------------------------------------------------------------------


@pytest.fixture
def sample_csv(tmp_path: Path) -> Path:
    p = tmp_path / "sample.csv"
    p.write_text("city,month,low,high\nMadrid,Jan,2,11\nBerlin,Jan,-3,3\n", encoding="utf-8")
    return p


@pytest.fixture
def sample_xlsx(tmp_path: Path) -> Path:
    path = tmp_path / "sample.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Data"
    ws.append(["city", "value"])
    ws.append(["Paris", 10])
    ws.append(["Lyon", 20])
    ws.append(["Total", 30])
    ws.merge_cells("A1:B1")
    wb.create_sheet("Other")
    wb.save(path)
    return path


# --------------------------------------------------------------------------
# semantic_types
# --------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("series", "name", "expected"),
    [
        (pd.Series([1, 2, 3, 4]), "value", "numeric"),
        (pd.Series([True, False, True]), "active", "boolean"),
        (pd.Series(["A", "B", "A", "B", "A"]), "group", "categorical"),
        # Regression: bare month names parse "successfully" via pd.to_datetime
        # against an implicit year/day and must not be misread as full dates.
        (pd.Series(["Jan", "Apr", "Jan", "Apr"]), "month", "categorical"),
        (pd.Series(["2024-01-15", "2024-02-20", "2024-03-05", "2024-04-10"]), "event_date", "datetime"),
        (pd.Series(["a@b.com", "c@d.org", "e@f.net", "g@h.io"]), "contact", "email"),
        (pd.Series(["https://a.com", "https://b.com", "http://c.org"]), "link", "url"),
        (pd.Series([48.85, 40.71, 51.50, -33.87]), "latitude", "latitude"),
        (pd.Series([2.35, -74.0, -0.12, 151.2]), "longitude", "longitude"),
        # Regression (TAB-03) : "eur"/"usd"/"gbp" doivent être délimités par
        # (^|_)...($|_), pas matchés en sous-chaîne libre -- "valeur" contient
        # "eur" mais n'est pas une colonne devise (bug réel : chart-type
        # sélectionné via ecdf au lieu de line, la colonne étant classée
        # "currency" au lieu de "numeric").
        (pd.Series([1000.0, 1200.0, 1500.0, 1800.0]), "valeur", "numeric"),
        (pd.Series([1000.0, 1200.0, 1500.0, 1800.0]), "montant_eur", "currency"),
        (pd.Series([f"user_{i}" for i in range(50)]), "user_id", "identifier"),
        (
            pd.Series([f"this is a fairly unique free-text sentence number {i}" for i in range(50)]),
            "comment",
            "text",
        ),
    ],
    ids=[
        "numeric",
        "boolean",
        "categorical_strings",
        "month_abbrev_not_datetime",
        "real_dates_datetime",
        "email",
        "url",
        "latitude_by_name_and_range",
        "longitude_by_name_and_range",
        "french_word_containing_eur_is_not_currency",
        "delimited_eur_suffix_is_currency",
        "identifier_by_name_and_uniqueness",
        "free_text_high_cardinality",
    ],
)
def test_detect_semantic_type(series: pd.Series, name: str, expected: str) -> None:
    assert detect_semantic_type(series, name=name) == expected


# --------------------------------------------------------------------------
# csv_reader
# --------------------------------------------------------------------------


def test_csv_read_workflow(sample_csv: Path) -> None:
    """sniff -> read -> preview: a CSV is detected, round-tripped, and truncated."""
    opts = sniff_csv(sample_csv)
    assert opts.delimiter == ","
    assert opts.header_row == 0

    df = read_csv(sample_csv)
    assert list(df.columns) == ["city", "month", "low", "high"]
    assert len(df) == 2

    assert len(preview_csv(sample_csv, n_rows=1)) == 1


def test_csv_fingerprint_is_stable_and_content_sensitive(sample_csv: Path, tmp_path: Path) -> None:
    other = tmp_path / "other.csv"
    other.write_text("a,b\n1,2\n", encoding="utf-8")
    assert csv_fingerprint(sample_csv) == csv_fingerprint(sample_csv)
    assert csv_fingerprint(sample_csv) != csv_fingerprint(other)


def test_csv_upload_guards(sample_csv: Path, tmp_path: Path) -> None:
    """Pre-read guards: empty frame and oversized upload both error; a normal
    file under the limit passes clean."""
    empty = csv_warnings(pd.DataFrame())
    assert empty and empty[0].severity == "error"

    big = tmp_path / "big.csv"
    big.write_bytes(b"x" * 1024)
    oversized = validate_upload_size(big, max_bytes=100)
    assert oversized and oversized[0].severity == "error"

    assert validate_upload_size(sample_csv, max_bytes=MAX_UPLOAD_BYTES) == []


# --------------------------------------------------------------------------
# excel_reader
# --------------------------------------------------------------------------


def test_excel_workflow(sample_xlsx: Path) -> None:
    """list sheets -> read named sheet -> structural warnings -> stable fingerprint."""
    assert list_sheets(sample_xlsx) == ["Data", "Other"]

    df = read_excel(sample_xlsx, sheet_name="Data")
    assert len(df) == 3

    messages = " ".join(w.message for w in excel_warnings(sample_xlsx, sheet_name="Data"))
    assert "merged cell" in messages
    assert "totals row" in messages

    assert excel_fingerprint(sample_xlsx) == excel_fingerprint(sample_xlsx)


# --------------------------------------------------------------------------
# clipboard
# --------------------------------------------------------------------------


@pytest.mark.parametrize(
    ("text", "expected_columns"),
    [
        ("city\tvalue\nParis\t10\nLyon\t20", ["city", "value"]),
        ("city,value\nParis,10\nLyon,20", ["city", "value"]),
        ("", []),
    ],
    ids=["tab_separated", "comma_separated", "empty"],
)
def test_parse_clipboard_text(text: str, expected_columns: list[str]) -> None:
    df = parse_clipboard_text(text)
    assert list(df.columns) == expected_columns
    if not expected_columns:
        assert df.empty


def test_clipboard_warnings_and_fingerprint() -> None:
    empty = clipboard_warnings(parse_clipboard_text(""))
    assert empty and empty[0].severity == "error"

    single = clipboard_warnings(parse_clipboard_text("just_one_column\nvalue1\nvalue2"))
    assert any("one column" in i.message for i in single)

    text = "city,value\nParis,10"
    assert clipboard_fingerprint(text) == clipboard_fingerprint(text)


# --------------------------------------------------------------------------
# profiler
# --------------------------------------------------------------------------


def test_profile_dataframe_workflow(sample_csv: Path) -> None:
    """read CSV -> profile: shape, column identity, and numeric column stats."""
    df = read_csv(sample_csv)
    profile = profile_dataframe(
        df, dataset_id="d1", fingerprint=csv_fingerprint(sample_csv), source_name="sample.csv"
    )
    assert profile.row_count == 2
    assert profile.column_count == 4
    assert {c.name for c in profile.columns} == {"city", "month", "low", "high"}

    low = profile.column("low")
    assert low is not None
    assert low.semantic_type == "numeric"
    assert low.minimum == -3.0
    assert low.maximum == 2.0


def test_profile_dataframe_structural_warnings() -> None:
    """Duplicate column names and entirely-empty columns are surfaced as warnings."""
    dup = profile_dataframe(
        pd.DataFrame([[1, 2]], columns=["x", "x"]), dataset_id="d1", fingerprint="f", source_name="dup.csv"
    )
    assert any("duplicate column name" in w.message for w in dup.warnings)

    empty_col = profile_dataframe(
        pd.DataFrame({"a": [1, 2, 3], "b": [None, None, None]}),
        dataset_id="d1",
        fingerprint="f",
        source_name="empty_col.csv",
    )
    assert any("entirely empty" in w.message and w.column == "b" for w in empty_col.warnings)


def test_profile_dataframe_data_quality_warnings() -> None:
    """Regression (TAB-05): a null value, an outlier, and a duplicate row must
    each surface as a warning -- previously `warnings` stayed empty despite
    `null_count`/`quantiles` already holding everything needed to detect them.
    """
    df = pd.DataFrame(
        {
            "amount": [10.0, 11.0, 12.0, 10.5, 11.5, 9.5, 10.0, 5000.0, None],
            "region": ["North", "South", "East", "West", "North", "South", "East", "West", "North"],
        }
    )
    # Duplicate the first row so a duplicate row is present.
    df = pd.concat([df, df.iloc[[0]]], ignore_index=True)

    profile = profile_dataframe(df, dataset_id="d1", fingerprint="f", source_name="anomalies.csv")

    assert any("manquante" in w.message and w.column == "amount" for w in profile.warnings), profile.warnings
    assert any("aberrante" in w.message and w.column == "amount" for w in profile.warnings), profile.warnings
    assert any("double" in w.message and w.column is None for w in profile.warnings), profile.warnings


def test_profile_dataframe_clean_data_has_no_data_quality_warnings() -> None:
    """A dataset with no nulls, no outliers, and no duplicate rows must not
    manufacture false positives."""
    df = pd.DataFrame({"amount": [10.0, 11.0, 12.0, 10.5, 11.5], "region": ["N", "S", "E", "W", "N"]})
    profile = profile_dataframe(df, dataset_id="d1", fingerprint="f", source_name="clean.csv")
    assert profile.warnings == []


def test_profile_dataframe_recognizes_psycopg_decimal_columns_as_numeric() -> None:
    """Regression: rows built from real psycopg SQL results (a raw list of
    dicts, `pd.DataFrame(rows)`, not `read_csv`) keep `decimal.Decimal`
    values in an `object`-dtype column -- `is_numeric_dtype` used to report
    False on that column even though every value is a number, so a monthly
    revenue total was classified "text"/"categorical" instead of "numeric",
    which cascaded into a wrong chart-recommendation goal and a column
    binding that plotted the wrong column entirely."""
    from decimal import Decimal

    rows = [{"month": m, "total_revenue": Decimal(str(200000 + m * 1000))} for m in range(1, 13)]
    df = pd.DataFrame(rows)
    profile = profile_dataframe(df, dataset_id="d1", fingerprint="f", source_name="psycopg.json")
    col = profile.column("total_revenue")
    assert col is not None
    assert col.physical_dtype == "float64"
    assert col.semantic_type in ("numeric", "currency")
    assert col.minimum == 201000.0
    assert col.maximum == 212000.0


def test_profile_dataframe_recognizes_psycopg_date_columns_as_datetime() -> None:
    """Regression: a psycopg SQL `DATE`/`TIMESTAMP` result column holds raw
    `datetime.date` objects in an `object`-dtype column when the frame is
    built from a list of row dicts -- `is_datetime64_any_dtype` used to
    report False, and `min()`/`max()` on that column crashed `ColumnProfile`
    construction outright (Pydantic expects `float | str`, got a bare
    `datetime.date`)."""
    from datetime import date

    rows = [{"period": date(2024, m, 1), "value": float(m)} for m in range(1, 13)]
    df = pd.DataFrame(rows)
    profile = profile_dataframe(df, dataset_id="d1", fingerprint="f", source_name="psycopg.json")
    col = profile.column("period")
    assert col is not None
    assert col.physical_dtype.startswith("datetime64")
    assert col.semantic_type == "datetime"
    # pandas coerces `date` to `Timestamp` (datetime precision): isoformat()
    # keeps the "T00:00:00" time component even for a date-only source value.
    assert col.minimum == pd.Timestamp(date(2024, 1, 1)).isoformat()
    assert col.maximum == pd.Timestamp(date(2024, 12, 1)).isoformat()


def test_profile_dataframe_leaves_genuine_text_columns_alone() -> None:
    """The Decimal/date coercion must not touch a column that is actually
    plain text just because every value happens to look number-like as a
    string -- only real Decimal/date instances trigger the coercion."""
    df = pd.DataFrame({"code": ["1", "2", "3"], "label": ["North", "South", "East"]})
    profile = profile_dataframe(df, dataset_id="d1", fingerprint="f", source_name="text.csv")
    code = profile.column("code")
    assert code is not None
    assert code.physical_dtype == "object"
    assert code.semantic_type in ("categorical", "text")
