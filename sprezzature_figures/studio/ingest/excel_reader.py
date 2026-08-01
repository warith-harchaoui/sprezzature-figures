"""
XLSX ingestion: sheet selection, header row choice, and warnings about
merged cells / unnamed columns / suspected total rows before the user
commits (plan §5.2). XLS (the legacy binary format) is not supported.

Author
------
Warith Harchaoui <warith.harchaoui@gmail.com>
"""

from __future__ import annotations

from pathlib import Path

import openpyxl
import pandas as pd
from os_helper import hashfile

from sprezzature_figures.core.dataset import DataWarning

_TOTAL_ROW_LABELS = {"total", "totals", "sum", "grand total"}


def list_sheets(path: str | Path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def read_excel(path: str | Path, *, sheet_name: str | int = 0, header_row: int = 0) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name, header=header_row, engine="openpyxl")


def preview_excel(path: str | Path, *, sheet_name: str | int = 0, header_row: int = 0, n_rows: int = 500) -> pd.DataFrame:
    return pd.read_excel(path, sheet_name=sheet_name, header=header_row, engine="openpyxl", nrows=n_rows)


def excel_fingerprint(path: str | Path) -> str:
    return hashfile(str(path))


def excel_warnings(path: str | Path, *, sheet_name: str | int = 0) -> list[DataWarning]:
    """Structural issues openpyxl can see that pandas silently flattens
    away: merged cells, and a last row that looks like a totals row a naive
    aggregation would otherwise double-count.
    """
    warnings: list[DataWarning] = []
    wb = openpyxl.load_workbook(path, read_only=False, data_only=True)
    try:
        ws = wb[sheet_name] if isinstance(sheet_name, str) else wb.worksheets[sheet_name]

        if ws.merged_cells.ranges:
            warnings.append(
                DataWarning(
                    message=f"sheet {ws.title!r} has {len(ws.merged_cells.ranges)} merged cell range(s)",
                    severity="warning",
                )
            )

        header = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(max_row=1))]
        if any(h == "" for h in header):
            warnings.append(
                DataWarning(message=f"sheet {ws.title!r} has one or more unnamed columns", severity="warning")
            )

        rows = list(ws.iter_rows(min_row=2, values_only=True))
        if rows:
            first_cell = rows[-1][0]
            if isinstance(first_cell, str) and first_cell.strip().lower() in _TOTAL_ROW_LABELS:
                warnings.append(
                    DataWarning(
                        message=f"sheet {ws.title!r} last row looks like a totals row ({first_cell!r})",
                        severity="warning",
                    )
                )
    finally:
        wb.close()

    return warnings
